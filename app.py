"""
테라로사 자사몰 주문취합 Streamlit 앱
실행: streamlit run app.py
"""

import json
import re
import os
import sqlite3
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook

# GitHub 영구 저장소
try:
    from github_storage import gh_load, gh_save
    _USE_GITHUB = True
except Exception:
    _USE_GITHUB = False
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────
# 설정 파일 경로
# ──────────────────────────────────────────────
CONFIG_PATH = Path("set_config.json")

# ──────────────────────────────────────────────
# 상품코드 DB (SQLite)
# ──────────────────────────────────────────────
DB_PATH = Path("product_codes.db")

def init_db():
    con = sqlite3.connect(DB_PATH)
    con.execute("""
        CREATE TABLE IF NOT EXISTS product_codes (
            code TEXT,
            name TEXT,
            option TEXT
        )
    """)
    con.execute("""
        CREATE TABLE IF NOT EXISTS db_meta (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    con.commit()
    con.close()

def save_codes_to_db(df: pd.DataFrame):
    """DataFrame → DB 저장 (기존 데이터 교체)"""
    con = sqlite3.connect(DB_PATH)
    con.execute("DELETE FROM product_codes")
    df.iloc[:, :3].to_sql("product_codes", con, if_exists="append", index=False)
    con.execute("INSERT OR REPLACE INTO db_meta VALUES ('updated_at', ?)",
                (datetime.now().strftime("%Y-%m-%d %H:%M"),))
    con.execute("INSERT OR REPLACE INTO db_meta VALUES ('row_count', ?)",
                (str(len(df)),))
    con.commit()
    con.close()

def load_codes_from_db() -> pd.DataFrame:
    """DB → DataFrame"""
    if not DB_PATH.exists():
        return pd.DataFrame()
    con = sqlite3.connect(DB_PATH)
    df = pd.read_sql("SELECT * FROM product_codes", con)
    con.close()
    df.fillna("", inplace=True)
    return df

def get_db_meta() -> dict:
    if not DB_PATH.exists():
        return {}
    con = sqlite3.connect(DB_PATH)
    rows = con.execute("SELECT key, value FROM db_meta").fetchall()
    con.close()
    return dict(rows)

# ──────────────────────────────────────────────
# 상수
# ──────────────────────────────────────────────
REMOVE_STRINGS = [
    "/구매 안함", "/플러스", "[플러스] ", "/불필요", "/필요",
    "불필요", "필요", "/상자 없음", "/포장 없음",
    "테라로사 시그니처 ", "[Online Exclusive] ", "[Online Exclusive/플러스] ",
    "[C.O.E/플러스] ",
]
TEXT_REPLACE = {
    "중간 분쇄(드립용)": "드립용",
    "가는 분쇄(에스프레소용)": "에스프레소용",
}
COLOR_DRIP   = "E2EFDA"
COLOR_SCOOP  = "FFF2CC"
COLOR_HEADER = "D9D9D9"
COLOR_WHITE  = "FFFFFF"
COLOR_SET    = "DDEEFF"   # 세트 분리 행 강조색
COL_WIDTHS   = {"A": 42, "B": 10, "C": 35, "D": 8, "E": 18}
THIN_BORDER  = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
WEIGHT_PATTERN = re.compile(r"(\d+(?:\.\d+)?\s*(?:kg|g))", re.IGNORECASE)

# ──────────────────────────────────────────────
# 세트 구성 설정 로드/저장
# ──────────────────────────────────────────────
def load_set_config() -> dict:
    if _USE_GITHUB:
        return gh_load("set_config.json", {})
    if CONFIG_PATH.exists():
        try:
            return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def save_set_config(cfg: dict):
    if _USE_GITHUB:
        gh_save("set_config.json", cfg)
    else:
        CONFIG_PATH.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")

# ──────────────────────────────────────────────
# 세트 분리 로직
# ──────────────────────────────────────────────
def expand_set_items(df: pd.DataFrame, set_config: dict) -> pd.DataFrame:
    """세트 상품을 구성 품목별 행으로 분리"""
    if not set_config:
        return df

    expanded = []
    for _, row in df.iterrows():
        name = str(row.get("품목명", "")).strip()
        matched = None
        for set_name, components in set_config.items():
            if set_name in name:
                matched = (set_name, components)
                break
        if matched and matched[1]:
            set_name, components = matched
            qty = int(row.get("수량", 1))
            for comp in components:
                new_row = row.copy()
                new_row["품목명_원본"] = f"[세트분리] {set_name} → {comp['name']}"
                # A열(품목명)은 세트명 그대로 유지, C열(옵션)에 구성 품목명 입력
                new_row["품목명"]  = set_name
                new_row["중량"]    = comp.get("weight", "")
                new_row["옵션"]    = comp["name"]
                new_row["수량"]    = qty * comp.get("qty", 1)
                new_row["_is_set_expanded"] = True
                expanded.append(new_row)
        else:
            row = row.copy()
            if "_is_set_expanded" not in row.index:
                row["_is_set_expanded"] = False
            expanded.append(row)
    return pd.DataFrame(expanded).reset_index(drop=True)

# ──────────────────────────────────────────────
# 무료원두 쿠폰 치환 로직
# ──────────────────────────────────────────────
def load_coupon_config() -> list:
    if _USE_GITHUB:
        return gh_load("coupon_config.json", [])
    coupon_path = Path("coupon_config.json")
    if coupon_path.exists():
        try:
            return json.loads(coupon_path.read_text(encoding="utf-8"))
        except Exception:
            pass
    return []

def expand_coupon_items(df: pd.DataFrame, coupon_config: list) -> pd.DataFrame:
    """무료원두 쿠폰 행을 설정한 품목들로 교체"""
    if not coupon_config:
        return df
    expanded = []
    for _, row in df.iterrows():
        name = str(row.get("품목명", "")).strip()
        if "무료원두 쿠폰" in name:
            for comp in coupon_config:
                if not comp.get("name"):
                    continue
                new_row = row.copy()
                new_row["품목명"]  = comp["name"]
                new_row["중량"]    = comp.get("weight", "250g")
                new_row["옵션"]    = comp.get("option", "증정 원두")
                new_row["수량"]    = int(comp.get("qty", 1))
                expanded.append(new_row)
        else:
            expanded.append(row)
    return pd.DataFrame(expanded).reset_index(drop=True)

# ──────────────────────────────────────────────
# 기존 처리 로직 (원본 스크립트 그대로)
# ──────────────────────────────────────────────
def load_order_data(file) -> pd.DataFrame:
    df = pd.read_excel(file, sheet_name="취합용", header=0, dtype=str)
    df = df.iloc[:, :2].copy()
    df.columns = ["품목명_원본", "수량"]
    df.dropna(subset=["품목명_원본"], inplace=True)
    df["수량"] = pd.to_numeric(df["수량"], errors="coerce").fillna(0).astype(int)
    return df.reset_index(drop=True)

def load_code_data(file) -> pd.DataFrame:
    df = pd.read_excel(file, header=0, dtype=str)
    df.fillna("", inplace=True)
    return df

def clean_item_name(name: str) -> str:
    for s in REMOVE_STRINGS:
        name = name.replace(s, "")
    for old, new in TEXT_REPLACE.items():
        name = name.replace(old, new)
    if "어센틱" in name and "정기 배송" in name:
        if "_" in name:
            option = name.split("_", 1)[1]
            name = "어센틱 에스프레소 블렌드_" + option
        else:
            name = "어센틱 에스프레소 블렌드"
    return name.strip()

def apply_sos_weight(name: str, weight: str) -> str:
    if "S.O.S" in name and weight == "300g":
        return "150g"
    return weight

def extract_weight(text: str):
    m = WEIGHT_PATTERN.search(text)
    if m:
        w = m.group(1).replace(" ", "")
        rest = (text[:m.start()] + text[m.end():]).strip().strip("/").strip()
        return w, rest
    return "", text
    
def split_item(raw_name: str):
    if "[커피 페스타 1+1]" in raw_name and ("KING콩" in raw_name or "King콩" in raw_name):
        item_part = raw_name.split("_", 1)[0].strip() if "_" in raw_name else raw_name
        item_part = re.sub(r"\+(블렌드|싱글오리진)$", "", item_part).strip()

        opt_raw = raw_name.split("_", 1)[1] if "_" in raw_name else ""
        for old, new in TEXT_REPLACE.items():
            opt_raw = opt_raw.replace(old, new)
        if "/" in opt_raw:
            grind_opt, gift_bean = opt_raw.split("/", 1)
            grind_opt = grind_opt.strip()
            gift_bean = gift_bean.strip()
        else:
            grind_opt = opt_raw.strip()
            gift_bean = ""
        row1 = (item_part, "250g", grind_opt)
        if gift_bean:
            return [row1, ("[커피 페스타 증정] " + gift_bean, "250g", "갈지않음")]
        return row1

    if "[커피 페스타 1+1]" in raw_name and "액상커피" in raw_name:
        item_name_orig = "[커피 페스타 1+1] 액상커피+파우더스틱"
        item_name_gift = "[커피 페스타 증정] 액상커피+파우더스틱"
        # _로 분리: 첫 번째 _ 뒤가 옵션
        parts = raw_name.split("_")
        opt_part = parts[1].strip() if len(parts) >= 2 else ""
        opt_part = re.sub(r"\(\d+개입\)", "", opt_part).strip()
        # 괄호 밖의 첫 번째 "+" 기준으로 분리
        depth, plus_idx = 0, -1
        for i, ch in enumerate(opt_part):
            if ch == "(": depth += 1
            elif ch == ")": depth -= 1
            elif ch == "+" and depth == 0:
                plus_idx = i; break
        if plus_idx != -1:
            opt1 = opt_part[:plus_idx].strip()
            opt2 = opt_part[plus_idx+1:].strip()
        else:
            opt1, opt2 = opt_part, ""
        row1 = (item_name_orig, "", opt1)
        if opt2:
            return [row1, (item_name_gift, "", opt2)]
        return row1

    if "[첫 구매 찬스]" in raw_name:
        name = raw_name.replace("[첫 구매 찬스] ", "").replace("250g", "").strip()
        return name, "250g", ""

    if "무료원두 쿠폰" in raw_name:
        return "무료원두 쿠폰 250g", "250g", "증정 원두"

    if "이 달의 킹콩" in raw_name \
            or "이달의 킹콩" in raw_name:
        return [
            ("[커피 페스타 1+1] 6월 KING콩 브라질 산투안토니우 엔리케",
             "250g", "플러스쿠폰"),
            ("[커피 페스타 1+1] 6월 KING콩 에티오피아 시다마 벤사",
             "250g", "플러스쿠폰"),
        ]

    if "이 달의 드립백" in raw_name or "이달의 드립백" in raw_name:
        if "_" in raw_name:
            parts = raw_name.split("_", 1)
            item_name = parts[0].strip()
            option = parts[1].strip()
            weight, option = extract_weight(option)
            return item_name, weight, option
        return raw_name, "", ""

    if "원두&커피 스쿱 세트" in raw_name or "원두 & 커피 스쿱 세트" in raw_name:
        if "_" in raw_name:
            parts = raw_name.split("_", 1)
            item_name = parts[0].strip()
            option = parts[1].strip().replace("갈지않음/", "").replace("(250g)", "").strip()
            return item_name, "250g", option
        return raw_name, "250g", ""

    if "_" in raw_name:
        parts = raw_name.split("_", 1)
        item_name = parts[0].strip()
        option = parts[1].strip()
        weight, option = extract_weight(option)
        return item_name, weight, option

    weight, rest = extract_weight(raw_name)
    if weight:
        return rest if rest else raw_name, weight, ""
    return raw_name, "", ""

def resolve_kingkong_name(df):
    # 이달의 킹콩은 split_item에서
    # 브라질/에티오피아 2행으로 분리 처리됨
    return df

def clean_kingkong_options(df: pd.DataFrame) -> pd.DataFrame:
    mask = df["품목명"].str.contains(r"[Kk][Ii][Nn][Gg]콩", na=False)
    for keyword in ["테라로사 바리스타", "에티오피아 농부", "멕시코 농장주"]:
        df.loc[mask, "옵션"] = df.loc[mask, "옵션"].str.replace(
            r"\s*/{1,2}\s*" + keyword, "", regex=True
        ).str.strip()
    df.loc[mask, "옵션"] = df.loc[mask, "옵션"].str.strip("/").str.strip()
    return df

def merge_gratitude_month(df: pd.DataFrame) -> pd.DataFrame:
    mask = df["품목명"].str.contains("감사의 달", na=False)
    if not mask.any():
        return df
    gdf, others = df[mask].copy(), df[~mask].copy()
    merged_rows = []
    gdf_with = gdf[gdf["옵션"].str.strip() != ""].copy()
    gdf_no   = gdf[gdf["옵션"].str.strip() == ""].copy()
    if not gdf_with.empty:
        gdf_with["_key"] = gdf_with["옵션"].str[:5]
        for key, g in gdf_with.groupby("_key", sort=False):
            merged_rows.append({"품목명": "[감사의 달] 2026 선물대전", "중량": g.iloc[0]["중량"],
                                 "옵션": min(g["옵션"].values, key=len), "수량": g["수량"].sum()})
    if not gdf_no.empty:
        for name, g in gdf_no.groupby("품목명", sort=False):
            merged_rows.append({"품목명": name, "중량": g.iloc[0]["중량"], "옵션": "", "수량": g["수량"].sum()})
    return pd.concat([others, pd.DataFrame(merged_rows)], ignore_index=True)

def classify(row) -> str:
    name, weight = row["품목명"], str(row["중량"])
    if "드립백" in name: return "드립백"
    if "원두&커피 스쿱 세트" in name or "원두 & 커피 스쿱 세트" in name: return "스쿱세트"
    if "세트" in name: return "세트"
    if re.search(r"\d+\s*(?:g|kg)", weight, re.IGNORECASE): return "원두"
    return "기타"

GROUP_ORDER = {"세트": 0, "기타": 1, "드립백": 2, "스쿱세트": 3, "원두": 4}

def weight_to_gram(w: str) -> float:
    w = str(w).strip()
    m = re.match(r"([\d.]+)\s*(kg|g)", w, re.IGNORECASE)
    if not m: return 0
    val = float(m.group(1))
    return val * 1000 if m.group(2).lower() == "kg" else val

def aggregate_and_sort(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_group"] = df.apply(classify, axis=1)
    def agg_key(row):
        g = row["_group"]
        if g in ("세트", "드립백", "스쿱세트"):
            return (row["품목명"], row["옵션"])
        return (row["품목명"], row["중량"], row["옵션"])
    df["_key"] = df.apply(agg_key, axis=1)
    df = df.groupby("_key", sort=False).agg(
        품목명=("품목명", "first"), 중량=("중량", "first"), 옵션=("옵션", "first"),
        수량=("수량", "sum"), _group=("_group", "first"),
    ).reset_index(drop=True)
    df["_g_order"] = df["_group"].map(GROUP_ORDER)
    df["_w_gram"]  = df["중량"].apply(weight_to_gram)
    df.sort_values(["_g_order", "품목명", "_w_gram", "옵션"],
                   ascending=[True, True, False, True], inplace=True)
    return df.reset_index(drop=True)

def match_code(row, code_df: pd.DataFrame) -> str:
    name   = str(row["품목명"]).strip()
    weight = str(row["중량"]).strip()
    option = str(row["옵션"]).strip()
    c_code, c_name, c_opt = code_df.columns[0], code_df.columns[1], code_df.columns[2]
    def eq(col, val): return code_df[col].str.strip() == val.strip()

    if re.search(r"[Kk][Ii][Nn][Gg]콩", name):
        res = code_df[eq(c_name, name) & (code_df[c_opt].str.strip() == "500g")]
        if not res.empty: return str(res.iloc[0][c_code])
        res = code_df[eq(c_name, name)]
        if not res.empty: return str(res.iloc[0][c_code])

    if "S.O.S" in name:
        res = code_df[eq(c_name, name) & eq(c_opt, weight)]
        if not res.empty: return str(res.iloc[0][c_code])

    if "스쿱 세트" in name or "스쿱세트" in name:
        res = code_df[eq(c_opt, option + "(250g)")]
        if not res.empty: return str(res.iloc[0][c_code])

    if "TO-GO" in name or "to-go" in name.lower():
        opt_no_color = re.sub(r"/블랙|/투명|/화이트|/레드", "", option).strip()
        res = code_df[eq(c_name, name) & eq(c_opt, opt_no_color)]
        if not res.empty: return str(res.iloc[0][c_code])

    if weight:
        res = code_df[eq(c_name, name) & eq(c_opt, weight)]
        if not res.empty: return str(res.iloc[0][c_code])

    res = code_df[eq(c_name, name) & eq(c_opt, option)]
    if not res.empty: return str(res.iloc[0][c_code])

    sorted_opt = "+".join(sorted(option.split("+")))
    name_rows  = code_df[code_df[c_name].str.strip() == name]
    if not name_rows.empty:
        match = name_rows[name_rows[c_opt].apply(
            lambda x: "+".join(sorted(str(x).split("+"))) == sorted_opt)]
        if not match.empty: return str(match.iloc[0][c_code])

    res = code_df[(code_df[c_name].str.strip() == name) & (code_df[c_opt].str.strip() == "")]
    if not res.empty: return str(res.iloc[0][c_code])

    if "옥스포드" in name:
        res = code_df[eq(c_name, name) & eq(c_opt, option)]
        if not res.empty: return str(res.iloc[0][c_code])

    return ""

def build_sheet3(raw_df: pd.DataFrame) -> pd.DataFrame:
    targets = {"테라로사 바리스타": "/ 테라로사 바리스타",
               "에티오피아 농부": "/ 에티오피아 농부",
               "멕시코 농장주": "/ 멕시코 농장주"}
    rows = []
    for label, keyword in targets.items():
        mask = raw_df["품목명_원본"].str.contains(keyword, na=False)
        qty  = raw_df.loc[mask, "수량"].sum()
        if qty > 0:
            rows.append({"품목명": "옥스포드 피규어", "빈칸": "", "이름": label, "수량": qty})
    return pd.DataFrame(rows)
    
def build_sheet2(main_df: pd.DataFrame) -> pd.DataFrame:
    rows = {}
    king_mask = (
        main_df["품목명"].str.contains(r"\[커피 페스타 1\+1\]", regex=True, na=False) &
        main_df["품목명"].str.contains("KING콩|King콩", na=False)
    )
    for _, r in main_df[(main_df["_group"] == "원두") & ~king_mask].iterrows():

        name = re.sub(r"^\[커피 페스타 증정\]\s*", "", r["품목명"]).strip()
        rows[name] = rows.get(name, 0) + weight_to_gram(r["중량"]) * r["수량"]
    for _, r in main_df[main_df["_group"] == "스쿱세트"].iterrows():
        name = r["옵션"]
        rows[name] = rows.get(name, 0) + 250 * r["수량"]
    for _, r in main_df[king_mask].iterrows():
        name = re.sub(r"^\[커피 페스타 1\+1\]\s*", "", r["품목명"]).strip()
        rows[name] = rows.get(name, 0) + 250 * r["수량"]

    return pd.DataFrame([{"품목명": n, "중량(kg)": round(g / 1000, 3)} for n, g in rows.items()])

def apply_style(ws, df_with_groups: pd.DataFrame):
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    drip_fill   = PatternFill("solid", fgColor=COLOR_DRIP)
    scoop_fill  = PatternFill("solid", fgColor=COLOR_SCOOP)
    white_fill  = PatternFill("solid", fgColor=COLOR_WHITE)
    set_fill    = PatternFill("solid", fgColor=COLOR_SET)
    header_font = Font(name="Arial", size=10, bold=True)
    body_font   = Font(name="Arial", size=10)

    headers = ["품목명", "중량", "옵션", "수량", "자사몰상품코드"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = THIN_BORDER

    row_num, prev_name = 2, None
    for _, r in df_with_groups.iterrows():
        cur_name = r["품목명"]
        group    = r.get("_group", "")
        is_set   = r.get("_is_set_expanded", False)

        if prev_name is not None and cur_name != prev_name:
            for col_idx in range(1, 6):
                ws.cell(row=row_num, column=col_idx).border = THIN_BORDER
            row_num += 1

        if is_set:         fill = set_fill
        elif group == "드립백":  fill = drip_fill
        elif group == "스쿱세트": fill = scoop_fill
        else:              fill = white_fill

        values = [cur_name, r["중량"], r["옵션"], r["수량"], r.get("상품코드", "")]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = body_font; cell.fill = fill; cell.border = THIN_BORDER

        prev_name = cur_name
        row_num += 1

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

def write_simple_sheet(ws, df: pd.DataFrame, title_row: list):
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(name="Arial", size=10, bold=True)
    body_font   = Font(name="Arial", size=10)
    for col_idx, h in enumerate(title_row, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = THIN_BORDER
    for r_idx, row in df.iterrows():
        for col_idx, val in enumerate(row.values, 1):
            cell = ws.cell(row=r_idx + 2, column=col_idx, value=val)
            cell.font = body_font; cell.border = THIN_BORDER

def insert_sheet3_into_sheet1(wb: Workbook):
    ws1, ws3 = wb["주문취합"], wb["바리스타·농부·농장주"]
    max_row_3   = ws3.max_row
    insert_count = max_row_3 - 1
    if insert_count <= 0: return
    ws1.insert_rows(2, amount=insert_count + 1)
    for src_row_idx in range(2, max_row_3 + 1):
        dest_row_idx = src_row_idx
        for col_idx in range(1, ws3.max_column + 1):
            src_cell = ws3.cell(row=src_row_idx, column=col_idx)
            dst_cell = ws1.cell(row=dest_row_idx, column=col_idx)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font); dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border); dst_cell.alignment = copy(src_cell.alignment)
    blank_row = 2 + insert_count
    for col_idx in range(1, 6):
        ws1.cell(row=blank_row, column=col_idx).border = THIN_BORDER

def postprocess_festa_rows(ws):
    # ── 1순위: 커피 페스타 포함 행 하늘색 ──
    FILL_FESTA = PatternFill("solid", fgColor="DDEEFF")
    for row in ws.iter_rows():
        a_val = row[0].value
        if a_val and "커피 페스타" in str(a_val):
            for cell in row[:5]:
                cell.fill = FILL_FESTA

    # ── 2순위: 플러스쿠폰 행 녹색 (하늘색 위에 덮어씀) ──
    FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
    for row in ws.iter_rows():
        c_val = row[2].value
        if c_val and "플러스쿠폰" in str(c_val):
            for cell in row[:5]:
                cell.fill = FILL_GREEN

    def insert_blank(ws, row_idx):
        ws.insert_rows(row_idx)
        for col in range(1, 6):
            c = ws.cell(row=row_idx, column=col)
            c.value = None
            c.fill = FILL_BLANK
            c.border = THIN_BORDER

    def is_king_brazil(a):
        return a and "[커피 페스타 1+1]" in str(a) and "KING콩" in str(a) and "브라질" in str(a)

    def is_king_ethiopia(a):
        return a and "[커피 페스타 1+1]" in str(a) and "KING콩" in str(a) and "에티오피아" in str(a)

    def is_gift_bean(a, b):
        return a and "[커피 페스타 증정]" in str(a) and str(b) == "250g"

    def is_festa_1p1_liquid(a):
        return a and "[커피 페스타 1+1]" in str(a) and "액상" in str(a)

    def is_gift_liquid(a, b):
        return a and "[커피 페스타 증정]" in str(a) and str(b) != "250g"

    def is_festa_item(a):
        return a and "[커피 페스타]" in str(a)

    # ── 빈행 제거: 페스타 증정 원두 행 사이 빈행 제거 ──
    changed = True
    while changed:
        changed = False
        rows_data = list(ws.iter_rows(values_only=True))
        for i in range(1, len(rows_data)):
            a = rows_data[i][0]
            if a is not None:
                continue
            prev_a = rows_data[i-1][0]
            next_a = rows_data[i+1][0] if i+1 < len(rows_data) else None
            prev_b = rows_data[i-1][1]
            next_b = rows_data[i+1][1] if i+1 < len(rows_data) else None
            if is_gift_bean(prev_a, prev_b) and is_gift_bean(next_a, next_b):
                ws.delete_rows(i + 1)
                changed = True
                break

    # ── 빈행 삽입: 아래서 위 순서로 ──
    rows_data = list(ws.iter_rows(values_only=True))
    to_insert = []

    for i in range(1, len(rows_data)):
        prev_a = rows_data[i-1][0]
        curr_a = rows_data[i][0]
        prev_b = rows_data[i-1][1]
        curr_b = rows_data[i][1]
        if prev_a is None or curr_a is None:
            continue

        # 1) 브라질 King콩 → 에티오피아 King콩
        if is_king_brazil(prev_a) and is_king_ethiopia(curr_a):
            to_insert.append(i + 1)
        # 2) 에티오피아 King콩 → 증정 원두
        elif is_king_ethiopia(prev_a) and is_gift_bean(curr_a, curr_b):
            to_insert.append(i + 1)
        # 3) 증정 원두 → 1+1 액상
        elif is_gift_bean(prev_a, prev_b) and is_festa_1p1_liquid(curr_a):
            to_insert.append(i + 1)
        # 4) 1+1 액상 → 증정 액상
        elif is_festa_1p1_liquid(prev_a) and is_gift_liquid(curr_a, curr_b):
            to_insert.append(i + 1)
        # 5) 증정 액상 → 다음 품목
        elif is_gift_liquid(prev_a, prev_b) and not is_gift_liquid(curr_a, curr_b):
            to_insert.append(i + 1)
        # 6) [커피 페스타] 단독 품목 사이 (품목명 다를 때)
        elif is_festa_item(prev_a) and is_festa_item(curr_a) and str(prev_a) != str(curr_a):
            to_insert.append(i + 1)

    for idx in sorted(set(to_insert), reverse=True):
        insert_blank(ws, idx)

    # ── 하늘색 채우기: 커피 페스타 포함 행 ──
    for row in ws.iter_rows():
        a_val = row[0].value
        if a_val and "커피 페스타" in str(a_val):
            for cell in row[:5]:
                cell.fill = FILL_FESTA

# ──────────────────────────────────────────────
# 메인 처리
# ──────────────────────────────────────────────
def process(order_file, set_config: dict, code_file=None) -> BytesIO:
    raw_df  = load_order_data(order_file)
    # 상품코드: 업로드 파일 우선, 없으면 DB에서 로드
    if code_file is not None:
        code_df = load_code_data(code_file)
    else:
        code_df = load_codes_from_db()
        if code_df.empty:
            raise ValueError("상품코드 DB가 비어 있습니다. 사이드바에서 상품코드 파일을 먼저 등록해 주세요.")
    sheet3_df = build_sheet3(raw_df)

    raw_df["품목명_정리"] = raw_df["품목명_원본"].apply(clean_item_name)

    expanded_rows = []
    for _, row in raw_df.iterrows():
        result = split_item(row["품목명_정리"])
        if isinstance(result, list):
            for r in result:
                new_row = row.copy(); new_row["품목명"] = r[0]
                new_row["중량"] = r[1]; new_row["옵션"] = r[2]
                expanded_rows.append(new_row)
        else:
            new_row = row.copy(); new_row["품목명"] = result[0]
            new_row["중량"] = result[1]; new_row["옵션"] = result[2]
            expanded_rows.append(new_row)
    raw_df = pd.DataFrame(expanded_rows).reset_index(drop=True)

    raw_df["중량"] = raw_df.apply(lambda r: apply_sos_weight(r["품목명"], r["중량"]), axis=1)
    raw_df = resolve_kingkong_name(raw_df)
    raw_df = clean_kingkong_options(raw_df)
    raw_df = merge_gratitude_month(raw_df)

    # ★ 무료원두 쿠폰 치환 적용
    coupon_config = load_coupon_config()
    raw_df = expand_coupon_items(raw_df, coupon_config)

    # ★ 세트 분리 적용
    raw_df = expand_set_items(raw_df, set_config)

    main_df = aggregate_and_sort(raw_df)

    # _is_set_expanded 컬럼 집계 후 복원
    if "_is_set_expanded" in raw_df.columns:
        set_flags = raw_df.groupby(
            raw_df.apply(lambda r: (r["품목명"], r.get("중량",""), r.get("옵션","")), axis=1)
        )["_is_set_expanded"].first()
        def get_flag(r):
            try: return set_flags[(r["품목명"], r["중량"], r["옵션"])]
            except: return False
        main_df["_is_set_expanded"] = main_df.apply(get_flag, axis=1)

    main_df["상품코드"] = main_df.apply(lambda r: match_code(r, code_df), axis=1)
    sheet2_df = build_sheet2(main_df)

    wb  = Workbook()
    ws1 = wb.active; ws1.title = "주문취합"
    apply_style(ws1, main_df)
    ws2 = wb.create_sheet("원두 중량 합산")
    write_simple_sheet(ws2, sheet2_df, ["품목명", "중량(kg)"])
    ws3 = wb.create_sheet("바리스타·농부·농장주")
    write_simple_sheet(ws3, sheet3_df, ["품목명", "빈칸", "이름", "수량"])
    insert_sheet3_into_sheet1(wb)
    postprocess_festa_rows(wb["주문취합"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ──────────────────────────────────────────────
# Streamlit UI
# ──────────────────────────────────────────────
st.set_page_config(page_title="테라로사 주문취합", page_icon="☕", layout="wide")

# ─── CSS ───
st.markdown("""
<style>
[data-testid="stSidebar"] { background: #FAF3F0; }
[data-testid="stSidebar"] .block-container { padding-top: 1.5rem; }
h1 { color: #8B3A2A !important; }
h2, h3 { color: #8B3A2A !important; }
.stButton > button {
    background: #8B3A2A; color: white; border: none;
    border-radius: 6px; font-weight: 600;
}
.stButton > button:hover { background: #C4644A; color: white; }
.stDownloadButton > button {
    background: #8B3A2A; color: white; border: none;
    border-radius: 6px; font-weight: 600; width: 100%;
}
.set-card {
    background: white; border: 1px solid #EDE5DC;
    border-radius: 10px; padding: 14px 16px; margin-bottom: 12px;
}
.set-badge {
    background: #F0DDD7; color: #8B3A2A; border-radius: 4px;
    padding: 2px 8px; font-size: 12px; font-weight: 600;
}
.comp-row {
    background: #FAF3F0; border-radius: 6px;
    padding: 6px 10px; margin: 4px 0; font-size: 13px;
    display: flex; justify-content: space-between;
    color: #2C2C2C;
}
</style>
""", unsafe_allow_html=True)

# ─── 상태 초기화 ───
init_db()
if "set_config" not in st.session_state:
    st.session_state.set_config = load_set_config()
if "editing_set" not in st.session_state:
    st.session_state.editing_set = None

# ═══════════════════════════════════════════
# 사이드바 — 상품코드 DB 관리
# ═══════════════════════════════════════════
with st.sidebar:
    st.markdown("## 🗄️ 상품코드 DB 관리")
    meta = get_db_meta()
    if meta:
        st.caption(f"등록 상품 {meta.get('row_count','?')}개 · 업데이트: {meta.get('updated_at','?')}")
    else:
        st.caption("DB 미등록")

    uploaded_code = st.file_uploader("상품코드 Excel 업로드", type=["xlsx"],
                                      key="sidebar_code_upload",
                                      help="월 1~2회 업로드하면 DB에 저장됩니다")
    if uploaded_code:
        if st.button("💾 DB에 저장", use_container_width=True, key="btn_save_db"):
            try:
                df_code = load_code_data(uploaded_code)
                save_codes_to_db(df_code)
                st.success(f"✅ {len(df_code)}개 상품코드 저장 완료!")
                st.rerun()
            except Exception as e:
                st.error(f"저장 실패: {e}")

    if meta:
        # ── 검색 후 수정 ──
        with st.expander("🔍 검색 후 수정/삭제"):
            search_kw = st.text_input("상품명 또는 옵션 검색", placeholder="예: 에티오피아",
                                       key="code_search_kw")
            if search_kw.strip():
                all_df = load_codes_from_db()
                cols = all_df.columns.tolist()
                mask = all_df.apply(
                    lambda r: search_kw.strip() in str(r.iloc[1]) or search_kw.strip() in str(r.iloc[2]),
                    axis=1
                )
                found = all_df[mask].reset_index(drop=True)
                if found.empty:
                    st.info("검색 결과 없음")
                else:
                    st.caption(f"{len(found)}개 검색됨")
                    for i, row in found.iterrows():
                        with st.container():
                            st.markdown(f"**{row.iloc[1]}** / {row.iloc[2]}")
                            st.caption(f"코드: {row.iloc[0]}")
                            ec1, ec2 = st.columns(2)
                            with ec1:
                                if st.button("✏️ 수정", key=f"edit_code_{i}", use_container_width=True):
                                    st.session_state[f"editing_code_{i}"] = True
                            with ec2:
                                if st.button("🗑️ 삭제", key=f"del_code_{i}", use_container_width=True):
                                    con = sqlite3.connect(DB_PATH)
                                    con.execute(
                                        "DELETE FROM product_codes WHERE code=? AND name=? AND option=?",
                                        (row.iloc[0], row.iloc[1], row.iloc[2])
                                    )
                                    # row_count 갱신
                                    cnt = con.execute("SELECT COUNT(*) FROM product_codes").fetchone()[0]
                                    con.execute("INSERT OR REPLACE INTO db_meta VALUES ('row_count', ?)", (str(cnt),))
                                    con.commit(); con.close()
                                    st.success("삭제 완료")
                                    st.rerun()
                            if st.session_state.get(f"editing_code_{i}"):
                                new_code = st.text_input("상품코드", value=row.iloc[0], key=f"ncode_{i}")
                                new_name = st.text_input("상품명",   value=row.iloc[1], key=f"nname_{i}")
                                new_opt  = st.text_input("옵션",     value=row.iloc[2], key=f"nopt_{i}")
                                if st.button("저장", key=f"save_code_{i}", use_container_width=True):
                                    con = sqlite3.connect(DB_PATH)
                                    con.execute(
                                        "UPDATE product_codes SET code=?, name=?, option=? WHERE code=? AND name=? AND option=?",
                                        (new_code, new_name, new_opt, row.iloc[0], row.iloc[1], row.iloc[2])
                                    )
                                    con.commit(); con.close()
                                    st.session_state.pop(f"editing_code_{i}", None)
                                    st.success("수정 완료")
                                    st.rerun()
                            st.divider()

        # ── 전체 테이블 편집 ──
        with st.expander("📋 전체 테이블 편집"):
            all_df = load_codes_from_db()
            col_names = ["상품코드", "상품명", "옵션"]
            all_df.columns = col_names
            edited = st.data_editor(
                all_df,
                num_rows="dynamic",
                use_container_width=True,
                hide_index=True,
                key="code_table_editor",
            )
            if st.button("💾 변경사항 저장", key="btn_save_edited", use_container_width=True):
                edited.columns = ["code", "name", "option"]
                edited = edited.dropna(subset=["code", "name"]).reset_index(drop=True)
                con = sqlite3.connect(DB_PATH)
                con.execute("DELETE FROM product_codes")
                edited.to_sql("product_codes", con, if_exists="append", index=False)
                con.execute("INSERT OR REPLACE INTO db_meta VALUES ('row_count', ?)", (str(len(edited)),))
                con.execute("INSERT OR REPLACE INTO db_meta VALUES ('updated_at', ?)",
                            (datetime.now().strftime("%Y-%m-%d %H:%M"),))
                con.commit(); con.close()
                st.success(f"✅ {len(edited)}개 저장 완료")
                st.rerun()

    st.divider()

    st.markdown("## ☕ 세트 상품 관리")
    st.caption("세트 상품을 구성 품목별로 분리합니다.")

    st.divider()

    # ── 새 세트 추가 ──
    with st.expander("➕ 새 세트 상품 추가", expanded=False):
        new_set_name = st.text_input("세트 상품명", placeholder="예: 간편커피&유리머그 세트",
                                      key="new_set_name")
        if st.button("추가", key="btn_add_set", use_container_width=True):
            name = new_set_name.strip()
            if name and name not in st.session_state.set_config:
                st.session_state.set_config[name] = []
                save_set_config(st.session_state.set_config)
                st.session_state.editing_set = name
                st.rerun()
            elif name in st.session_state.set_config:
                st.warning("이미 등록된 세트 상품입니다.")
            else:
                st.warning("세트 상품명을 입력하세요.")

    st.divider()

    # ── 세트 목록 ──
    if not st.session_state.set_config:
        st.info("등록된 세트 상품이 없습니다.")
    else:
        st.markdown(f"**등록된 세트** {len(st.session_state.set_config)}개")
        for set_name in list(st.session_state.set_config.keys()):
            comps = st.session_state.set_config[set_name]
            is_editing = st.session_state.editing_set == set_name

            with st.container():
                col_name, col_edit, col_del = st.columns([6, 2, 2])
                with col_name:
                    st.markdown(f"**{set_name}**")
                    st.caption(f"구성 {len(comps)}개")
                with col_edit:
                    if st.button("편집" if not is_editing else "닫기",
                                  key=f"edit_{set_name}", use_container_width=True):
                        st.session_state.editing_set = None if is_editing else set_name
                        st.rerun()
                with col_del:
                    if st.button("삭제", key=f"del_{set_name}", use_container_width=True,
                                  type="secondary"):
                        del st.session_state.set_config[set_name]
                        if st.session_state.editing_set == set_name:
                            st.session_state.editing_set = None
                        save_set_config(st.session_state.set_config)
                        st.rerun()

                # ── 구성 옵션 편집 패널 ──
                if is_editing:
                    with st.container():
                        st.markdown(f"###### 구성 품목 — {set_name}")

                        # 기존 구성 목록
                        for i, comp in enumerate(comps):
                            c1, c2, c3, c4, c5 = st.columns([1, 4, 2, 2, 1])
                            with c1:
                                new_qty = st.number_input("수량", min_value=1,
                                    value=comp.get("qty", 1),
                                    key=f"qty_{set_name}_{i}", label_visibility="collapsed")
                            with c2:
                                new_name = st.text_input("품목명",
                                    value=comp.get("name", ""),
                                    key=f"cname_{set_name}_{i}", label_visibility="collapsed")
                            with c3:
                                new_weight = st.text_input("중량",
                                    value=comp.get("weight", ""),
                                    placeholder="예: 250g",
                                    key=f"cweight_{set_name}_{i}", label_visibility="collapsed")
                            with c4:
                                new_option = st.text_input("옵션",
                                    value=comp.get("option", ""),
                                    placeholder="옵션(선택)",
                                    key=f"coption_{set_name}_{i}", label_visibility="collapsed")
                            with c5:
                                if st.button("✕", key=f"rm_{set_name}_{i}"):
                                    comps.pop(i)
                                    save_set_config(st.session_state.set_config)
                                    st.rerun()

                            # 실시간 저장
                            comps[i] = {"name": new_name, "qty": int(new_qty),
                                         "weight": new_weight, "option": new_option}

                        # 새 구성 품목 추가
                        st.markdown("---")
                        na1, na2, na3, na4, na5 = st.columns([1, 4, 2, 2, 1])
                        with na1:
                            add_qty = st.number_input("수량", min_value=1, value=1,
                                key=f"addqty_{set_name}", label_visibility="collapsed")
                        with na2:
                            add_name = st.text_input("품목명", placeholder="구성 품목명",
                                key=f"addname_{set_name}", label_visibility="collapsed")
                        with na3:
                            add_weight = st.text_input("중량", placeholder="예: 250g",
                                key=f"addweight_{set_name}", label_visibility="collapsed")
                        with na4:
                            add_option = st.text_input("옵션", placeholder="옵션(선택)",
                                key=f"addoption_{set_name}", label_visibility="collapsed")
                        with na5:
                            if st.button("＋", key=f"addcomp_{set_name}"):
                                if add_name.strip():
                                    comps.append({"name": add_name.strip(),
                                                   "qty": int(add_qty),
                                                   "weight": add_weight.strip(),
                                                   "option": add_option.strip()})
                                    save_set_config(st.session_state.set_config)
                                    st.rerun()

                        # 저장 버튼
                        if st.button("💾 저장", key=f"save_{set_name}", use_container_width=True):
                            save_set_config(st.session_state.set_config)
                            st.success("저장 완료!")

                st.divider()

# ═══════════════════════════════════════════
# 메인 화면
# ═══════════════════════════════════════════
st.title("테라로사 자사몰 주문취합")

# ── 세트 구성 현황 요약 ──
if st.session_state.set_config:
    with st.expander(f"📦 세트 분리 설정 — {len(st.session_state.set_config)}개 등록됨", expanded=False):
        cols = st.columns(min(3, len(st.session_state.set_config)))
        for i, (sname, comps) in enumerate(st.session_state.set_config.items()):
            with cols[i % 3]:
                st.markdown(f"**{sname}**")
                for c in comps:
                    weight_str = f" {c['weight']}" if c.get("weight") else ""
                    option_str = f" / {c['option']}" if c.get("option") else ""
                    st.caption(f"× {c['qty']}  {c['name']}{weight_str}{option_str}")

st.divider()

# ── 파일 업로드 ──
col1, col2 = st.columns(2)
with col1:
    order_file = st.file_uploader("📄 주문취합 Excel", type=["xlsx"],
                                   help="'취합용' 시트가 포함된 주문 파일")
with col2:
    meta = get_db_meta()
    if meta:
        st.success(f"✅ 상품코드 DB 등록됨\n\n"
                   f"- 상품 수: **{meta.get('row_count', '?')}개**\n"
                   f"- 최종 업데이트: {meta.get('updated_at', '?')}")
    else:
        st.warning("⚠️ 상품코드 DB 없음\n\n사이드바에서 상품코드 파일을 등록해 주세요.")

st.divider()

# ── 처리 및 다운로드 ──
db_ready = DB_PATH.exists() and not load_codes_from_db().empty
if order_file and db_ready:
    if st.button("🚀 주문 취합 처리 시작", use_container_width=True):
        with st.spinner("처리 중..."):
            try:
                result_buf = process(order_file, st.session_state.set_config)
                today = datetime.today().strftime("%Y%m%d")
                st.success("✅ 처리 완료!")

                set_count = len(st.session_state.set_config)
                if set_count:
                    st.info(f"📦 세트 분리 적용: {set_count}개 세트 상품 → 구성 품목별 행으로 분리됨 (하늘색 강조)")

                st.download_button(
                    label="⬇️ 결과 Excel 다운로드",
                    data=result_buf,
                    file_name=f"자사몰주문취합_{today}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"❌ 오류 발생: {e}")
                st.exception(e)
elif not db_ready:
    st.info("👆 사이드바에서 상품코드 DB를 먼저 등록해 주세요.")
else:
    st.info("👆 주문취합 파일을 업로드하면 처리 버튼이 활성화됩니다.")

# ── 하단 안내 ──
with st.expander("ℹ️ 사용 방법"):
    st.markdown("""
**상품코드 DB 등록 (월 1~2회)**
1. 왼쪽 사이드바 상단 → **상품코드 DB 관리** 에서 `자사몰상품코드.xlsx` 업로드
2. **DB에 저장** 버튼 클릭 → 이후 매일 주문 처리 시 자동으로 DB 사용

**주문 취합 처리 (매일)**
1. 주문취합 Excel 업로드 → **주문 취합 처리 시작** 클릭
2. 결과 Excel 다운로드

**세트 상품 분리 설정**
1. 왼쪽 사이드바 → **세트 상품 관리** 에서 세트명 추가
2. 편집 버튼으로 구성 품목(품목명, 수량, 중량, 옵션) 입력 후 저장
3. 설정은 `set_config.json`에 자동 저장 → 다음 실행에도 유지

**처리 결과**
- 세트 분리된 행은 주문취합 시트에서 **하늘색**으로 표시됩니다
- 세트 1개 주문 × 구성 수량으로 자동 계산됩니다
""")
