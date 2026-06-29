"""
테라로사 세트 상품 관리 페이지
기존 app.py와 같은 저장소의 pages/ 폴더에 넣으면 사이드바 메뉴로 자동 등록됩니다.
"""

import json
import re
import os
from copy import copy
from datetime import datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ──────────────────────────────────────────────
# 설정 파일 경로
# ──────────────────────────────────────────────
CONFIG_PATH = Path("set_config.json")

# ──────────────────────────────────────────────
# 상수
# ──────────────────────────────────────────────
REMOVE_STRINGS = [
    "/구매 안함", "/플러스", "[플러스] ", "/불필요", "/필요",
    "불필요", "필요", "/상자 없음", "/포장 없음",
    "테라로사 시그니처 ", "[Online Exclusive] ", "[Online Exclusive/플러스] ",
    "[C.O.E/플러스] ",
]
TEXT_REPLACE = {
    "중간 분쇄(드립용)": "드립용",
    "가는 분쇄(에스프레소용)": "에스프레소용",
}
COLOR_DRIP   = "E2EFDA"
COLOR_SCOOP  = "FFF2CC"
COLOR_HEADER = "D9D9D9"
COLOR_WHITE  = "FFFFFF"
COLOR_SET    = "DDEEFF"
COL_WIDTHS   = {"A": 42, "B": 10, "C": 35, "D": 8, "E": 18}
THIN_BORDER  = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)
WEIGHT_PATTERN = re.compile(r"(\d+(?:\.\d+)?\s*(?:kg|g))", re.IGNORECASE)

# ──────────────────────────────────────────────
# 세트 구성 설정 로드/저장
# ──────────────────────────────────────────────
def load_set_config() -> dict:
    if CONFIG_PATH.exists():
        try:
            return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        except Exception:
            pass
    return {}

def save_set_config(cfg: dict):
    CONFIG_PATH.write_text(json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")

# ──────────────────────────────────────────────
# 세트 분리 로직
# ──────────────────────────────────────────────
def expand_set_items(df: pd.DataFrame, set_config: dict) -> pd.DataFrame:
    if not set_config:
        return df
    expanded = []
    for _, row in df.iterrows():
        name = str(row.get("품목명", "")).strip()
        matched = None
        for set_name, components in set_config.items():
            if set_name in name:
                matched = (set_name, components)
                break
        if matched and matched[1]:
            set_name, components = matched
            qty = int(row.get("수량", 1))
            for comp in components:
                new_row = row.copy()
                new_row["품목명_원본"] = f"[세트분리] {set_name} → {comp['name']}"
                new_row["품목명"]  = comp["name"]
                new_row["중량"]    = comp.get("weight", "")
                new_row["옵션"]    = comp.get("option", "")
                new_row["수량"]    = qty * comp.get("qty", 1)
                new_row["_is_set_expanded"] = True
                expanded.append(new_row)
        else:
            row = row.copy()
            if "_is_set_expanded" not in row.index:
                row["_is_set_expanded"] = False
            expanded.append(row)
    return pd.DataFrame(expanded).reset_index(drop=True)

# ──────────────────────────────────────────────
# 기존 처리 로직
# ──────────────────────────────────────────────
def load_order_data(file) -> pd.DataFrame:
    df = pd.read_excel(file, sheet_name="취합용", header=0, dtype=str)
    df = df.iloc[:, :2].copy()
    df.columns = ["품목명_원본", "수량"]
    df.dropna(subset=["품목명_원본"], inplace=True)
    df["수량"] = pd.to_numeric(df["수량"], errors="coerce").fillna(0).astype(int)
    return df.reset_index(drop=True)

def load_code_data(file) -> pd.DataFrame:
    df = pd.read_excel(file, header=0, dtype=str)
    df.fillna("", inplace=True)
    return df

def clean_item_name(name: str) -> str:
    for s in REMOVE_STRINGS:
        name = name.replace(s, "")
    for old, new in TEXT_REPLACE.items():
        name = name.replace(old, new)
    if "어센틱" in name and "정기 배송" in name:
        if "_" in name:
            option = name.split("_", 1)[1]
            name = "어센틱 에스프레소 블렌드_" + option
        else:
            name = "어센틱 에스프레소 블렌드"
    return name.strip()

def apply_sos_weight(name: str, weight: str) -> str:
    if "S.O.S" in name and weight == "300g":
        return "150g"
    return weight

def extract_weight(text: str):
    m = WEIGHT_PATTERN.search(text)
    if m:
        w = m.group(1).replace(" ", "")
        rest = (text[:m.start()] + text[m.end():]).strip().strip("/").strip()
        return w, rest
    return "", text

def split_item(raw_name: str):
    if "[커피 페스타 1+1]" in raw_name and "King콩" in raw_name:
        item_part, opt_part = (raw_name.split("_", 1) if "_" in raw_name else (raw_name, ""))
        item_name = item_part.strip()
        opt_part = re.sub(r"\(\d*g?\)", "", opt_part).strip()
        if "/" in opt_part:
            before_slash, extra_bean = opt_part.split("/", 1)
            extra_bean = extra_bean.strip()
        else:
            before_slash, extra_bean = opt_part, ""
        opt1 = before_slash.strip().strip("/").strip()
        row1 = (item_name, "250g", opt1)
        if extra_bean:
            return [row1, ("[커피 페스타 증정] " + extra_bean, "250g", "갈지않음")]
        return row1

    if "[커피 페스타 1+1]" in raw_name and "액상커피" in raw_name:
        item_name_orig = "[커피 페스타 1+1] 액상커피+파우더스틱"
        item_name_gift = "[커피 페스타 증정] 액상커피+파우더스틱"
        parts = raw_name.split("_")
        opt_part = parts[1].strip() if len(parts) >= 2 else ""
        opt_part = re.sub(r"\(\d+개입\)", "", opt_part).strip()
        depth, plus_idx = 0, -1
        for i, ch in enumerate(opt_part):
            if ch == "(": depth += 1
            elif ch == ")": depth -= 1
            elif ch == "+" and depth == 0:
                plus_idx = i; break
        if plus_idx != -1:
            opt1 = opt_part[:plus_idx].strip()
            opt2 = opt_part[plus_idx+1:].strip()
        else:
            opt1, opt2 = opt_part, ""
        row1 = (item_name_orig, "", opt1)
        if opt2:
            return [row1, (item_name_gift, "", opt2)]
        return row1

    if "[첫 구매 찬스]" in raw_name:
        name = raw_name.replace("[첫 구매 찬스] ", "").replace("250g", "").strip()
        return name, "250g", ""

    if "무료원두 쿠폰" in raw_name:
        return "무료원두 쿠폰 250g", "250g", "증정 원두"

    if "이 달의 킹콩" in raw_name or "이달의 킹콩" in raw_name:
        return raw_name, "500g", "플러스쿠폰"

    if "이 달의 드립백" in raw_name or "이달의 드립백" in raw_name:
        if "_" in raw_name:
            parts = raw_name.split("_", 1)
            item_name = parts[0].strip()
            option = parts[1].strip()
            weight, option = extract_weight(option)
            return item_name, weight, option
        return raw_name, "", ""

    if "원두&커피 스쿱 세트" in raw_name or "원두 & 커피 스쿱 세트" in raw_name:
        if "_" in raw_name:
            parts = raw_name.split("_", 1)
            item_name = parts[0].strip()
            option = parts[1].strip().replace("갈지않음/", "").replace("(250g)", "").strip()
            return item_name, "250g", option
        return raw_name, "250g", ""

    if "_" in raw_name:
        parts = raw_name.split("_", 1)
        item_name = parts[0].strip()
        option = parts[1].strip()
        weight, option = extract_weight(option)
        return item_name, weight, option

    weight, rest = extract_weight(raw_name)
    if weight:
        return rest if rest else raw_name, weight, ""
    return raw_name, "", ""

def resolve_kingkong_name(df: pd.DataFrame) -> pd.DataFrame:
    king_rows = df[df["품목명"].str.contains(r"[Kk][Ii][Nn][Gg]콩", na=False)]
    mask = df["품목명"].str.contains("이 달의 킹콩|이달의 킹콩", na=False)
    if not mask.any():
        return df
    king_name = king_rows.iloc[0]["품목명"] if not king_rows.empty else f"[{datetime.today().month}월 KING콩]"
    df.loc[mask, "품목명"] = king_name
    df.loc[mask, "중량"]   = "500g"
    df.loc[mask, "옵션"]   = "플러스쿠폰"
    return df

def clean_kingkong_options(df: pd.DataFrame) -> pd.DataFrame:
    mask = df["품목명"].str.contains(r"[Kk][Ii][Nn][Gg]콩", na=False)
    for keyword in ["테라로사 바리스타", "에티오피아 농부", "멕시코 농장주"]:
        df.loc[mask, "옵션"] = df.loc[mask, "옵션"].str.replace(
            r"\s*/{1,2}\s*" + keyword, "", regex=True
        ).str.strip()
    df.loc[mask, "옵션"] = df.loc[mask, "옵션"].str.strip("/").str.strip()
    return df

def merge_gratitude_month(df: pd.DataFrame) -> pd.DataFrame:
    mask = df["품목명"].str.contains("감사의 달", na=False)
    if not mask.any():
        return df
    gdf, others = df[mask].copy(), df[~mask].copy()
    merged_rows = []
    gdf_with = gdf[gdf["옵션"].str.strip() != ""].copy()
    gdf_no   = gdf[gdf["옵션"].str.strip() == ""].copy()
    if not gdf_with.empty:
        gdf_with["_key"] = gdf_with["옵션"].str[:5]
        for key, g in gdf_with.groupby("_key", sort=False):
            merged_rows.append({"품목명": "[감사의 달] 2026 선물대전", "중량": g.iloc[0]["중량"],
                                 "옵션": min(g["옵션"].values, key=len), "수량": g["수량"].sum()})
    if not gdf_no.empty:
        for name, g in gdf_no.groupby("품목명", sort=False):
            merged_rows.append({"품목명": name, "중량": g.iloc[0]["중량"], "옵션": "", "수량": g["수량"].sum()})
    return pd.concat([others, pd.DataFrame(merged_rows)], ignore_index=True)

def classify(row) -> str:
    name, weight = row["품목명"], str(row["중량"])
    if "드립백" in name: return "드립백"
    if "원두&커피 스쿱 세트" in name or "원두 & 커피 스쿱 세트" in name: return "스쿱세트"
    if "세트" in name: return "세트"
    if re.search(r"\d+\s*(?:g|kg)", weight, re.IGNORECASE): return "원두"
    return "기타"

GROUP_ORDER = {"세트": 0, "기타": 1, "드립백": 2, "스쿱세트": 3, "원두": 4}

def weight_to_gram(w: str) -> float:
    w = str(w).strip()
    m = re.match(r"([\d.]+)\s*(kg|g)", w, re.IGNORECASE)
    if not m: return 0
    val = float(m.group(1))
    return val * 1000 if m.group(2).lower() == "kg" else val

def aggregate_and_sort(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df["_group"] = df.apply(classify, axis=1)
    def agg_key(row):
        g = row["_group"]
        if g in ("세트", "드립백", "스쿱세트"):
            return (row["품목명"], row["옵션"])
        return (row["품목명"], row["중량"], row["옵션"])
    df["_key"] = df.apply(agg_key, axis=1)
    df = df.groupby("_key", sort=False).agg(
        품목명=("품목명", "first"), 중량=("중량", "first"), 옵션=("옵션", "first"),
        수량=("수량", "sum"), _group=("_group", "first"),
    ).reset_index(drop=True)
    df["_g_order"] = df["_group"].map(GROUP_ORDER)
    df["_w_gram"]  = df["중량"].apply(weight_to_gram)
    df.sort_values(["_g_order", "품목명", "_w_gram", "옵션"],
                   ascending=[True, True, False, True], inplace=True)
    return df.reset_index(drop=True)

def match_code(row, code_df: pd.DataFrame) -> str:
    name   = str(row["품목명"]).strip()
    weight = str(row["중량"]).strip()
    option = str(row["옵션"]).strip()
    c_code, c_name, c_opt = code_df.columns[0], code_df.columns[1], code_df.columns[2]
    def eq(col, val): return code_df[col].str.strip() == val.strip()

    if re.search(r"[Kk][Ii][Nn][Gg]콩", name):
        res = code_df[eq(c_name, name) & (code_df[c_opt].str.strip() == "500g")]
        if not res.empty: return str(res.iloc[0][c_code])
        res = code_df[eq(c_name, name)]
        if not res.empty: return str(res.iloc[0][c_code])

    if "S.O.S" in name:
        res = code_df[eq(c_name, name) & eq(c_opt, weight)]
        if not res.empty: return str(res.iloc[0][c_code])

    if "스쿱 세트" in name or "스쿱세트" in name:
        res = code_df[eq(c_opt, option + "(250g)")]
        if not res.empty: return str(res.iloc[0][c_code])

    if "TO-GO" in name or "to-go" in name.lower():
        opt_no_color = re.sub(r"/블랙|/투명|/화이트|/레드", "", option).strip()
        res = code_df[eq(c_name, name) & eq(c_opt, opt_no_color)]
        if not res.empty: return str(res.iloc[0][c_code])

    if weight:
        res = code_df[eq(c_name, name) & eq(c_opt, weight)]
        if not res.empty: return str(res.iloc[0][c_code])

    res = code_df[eq(c_name, name) & eq(c_opt, option)]
    if not res.empty: return str(res.iloc[0][c_code])

    sorted_opt = "+".join(sorted(option.split("+")))
    name_rows  = code_df[code_df[c_name].str.strip() == name]
    if not name_rows.empty:
        match = name_rows[name_rows[c_opt].apply(
            lambda x: "+".join(sorted(str(x).split("+"))) == sorted_opt)]
        if not match.empty: return str(match.iloc[0][c_code])

    res = code_df[(code_df[c_name].str.strip() == name) & (code_df[c_opt].str.strip() == "")]
    if not res.empty: return str(res.iloc[0][c_code])

    if "옥스포드" in name:
        res = code_df[eq(c_name, name) & eq(c_opt, option)]
        if not res.empty: return str(res.iloc[0][c_code])

    return ""

def build_sheet3(raw_df: pd.DataFrame) -> pd.DataFrame:
    targets = {"테라로사 바리스타": "/ 테라로사 바리스타",
               "에티오피아 농부": "/ 에티오피아 농부",
               "멕시코 농장주": "/ 멕시코 농장주"}
    rows = []
    for label, keyword in targets.items():
        mask = raw_df["품목명_원본"].str.contains(keyword, na=False)
        qty  = raw_df.loc[mask, "수량"].sum()
        if qty > 0:
            rows.append({"품목명": "옥스포드 피규어", "빈칸": "", "이름": label, "수량": qty})
    return pd.DataFrame(rows)

def build_sheet2(main_df: pd.DataFrame) -> pd.DataFrame:
    rows = {}
    for _, r in main_df[main_df["_group"] == "원두"].iterrows():
        # [커피 페스타 증정] 등 접두어 제거 후 품목명 기준 합산
        name = re.sub(r"^\[커피 페스타 증정\]\s*", "", r["품목명"]).strip()
        rows[name] = rows.get(name, 0) + weight_to_gram(r["중량"]) * r["수량"]
    for _, r in main_df[main_df["_group"] == "스쿱세트"].iterrows():
        name = r["옵션"]
        rows[name] = rows.get(name, 0) + 250 * r["수량"]
    return pd.DataFrame([{"품목명": n, "중량(kg)": round(g / 1000, 3)} for n, g in rows.items()])

def apply_style(ws, df_with_groups: pd.DataFrame):
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    drip_fill   = PatternFill("solid", fgColor=COLOR_DRIP)
    scoop_fill  = PatternFill("solid", fgColor=COLOR_SCOOP)
    white_fill  = PatternFill("solid", fgColor=COLOR_WHITE)
    set_fill    = PatternFill("solid", fgColor=COLOR_SET)
    header_font = Font(name="Arial", size=10, bold=True)
    body_font   = Font(name="Arial", size=10)

    headers = ["품목명", "중량", "옵션", "수량", "자사몰상품코드"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = THIN_BORDER

    row_num, prev_name = 2, None
    for _, r in df_with_groups.iterrows():
        cur_name = r["품목명"]
        group    = r.get("_group", "")
        is_set   = r.get("_is_set_expanded", False)

        if prev_name is not None and cur_name != prev_name:
            for col_idx in range(1, 6):
                ws.cell(row=row_num, column=col_idx).border = THIN_BORDER
            row_num += 1

        if is_set:              fill = set_fill
        elif group == "드립백":  fill = drip_fill
        elif group == "스쿱세트": fill = scoop_fill
        else:                   fill = white_fill

        values = [cur_name, r["중량"], r["옵션"], r["수량"], r.get("상품코드", "")]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=val)
            cell.font = body_font; cell.fill = fill; cell.border = THIN_BORDER

        prev_name = cur_name
        row_num += 1

    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

def write_simple_sheet(ws, df: pd.DataFrame, title_row: list):
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER)
    header_font = Font(name="Arial", size=10, bold=True)
    body_font   = Font(name="Arial", size=10)
    for col_idx, h in enumerate(title_row, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center"); cell.border = THIN_BORDER
    for r_idx, row in df.iterrows():
        for col_idx, val in enumerate(row.values, 1):
            cell = ws.cell(row=r_idx + 2, column=col_idx, value=val)
            cell.font = body_font; cell.border = THIN_BORDER

def insert_sheet3_into_sheet1(wb: Workbook):
    ws1, ws3 = wb["주문취합"], wb["바리스타·농부·농장주"]
    max_row_3    = ws3.max_row
    insert_count = max_row_3 - 1
    if insert_count <= 0: return
    ws1.insert_rows(2, amount=insert_count + 1)
    for src_row_idx in range(2, max_row_3 + 1):
        for col_idx in range(1, ws3.max_column + 1):
            src_cell = ws3.cell(row=src_row_idx, column=col_idx)
            dst_cell = ws1.cell(row=src_row_idx, column=col_idx)
            dst_cell.value = src_cell.value
            if src_cell.has_style:
                dst_cell.font = copy(src_cell.font); dst_cell.fill = copy(src_cell.fill)
                dst_cell.border = copy(src_cell.border); dst_cell.alignment = copy(src_cell.alignment)
    blank_row = 2 + insert_count
    for col_idx in range(1, 6):
        ws1.cell(row=blank_row, column=col_idx).border = THIN_BORDER

def process(order_file, code_file, set_config: dict) -> BytesIO:
    raw_df  = load_order_data(order_file)
    code_df = load_code_data(code_file)
    sheet3_df = build_sheet3(raw_df)

    raw_df["품목명_정리"] = raw_df["품목명_원본"].apply(clean_item_name)

    expanded_rows = []
    for _, row in raw_df.iterrows():
        result = split_item(row["품목명_정리"])
        if isinstance(result, list):
            for r in result:
                new_row = row.copy(); new_row["품목명"] = r[0]
                new_row["중량"] = r[1]; new_row["옵션"] = r[2]
                expanded_rows.append(new_row)
        else:
            new_row = row.copy(); new_row["품목명"] = result[0]
            new_row["중량"] = result[1]; new_row["옵션"] = result[2]
            expanded_rows.append(new_row)
    raw_df = pd.DataFrame(expanded_rows).reset_index(drop=True)

    raw_df["중량"] = raw_df.apply(lambda r: apply_sos_weight(r["품목명"], r["중량"]), axis=1)
    raw_df = resolve_kingkong_name(raw_df)
    raw_df = clean_kingkong_options(raw_df)
    raw_df = merge_gratitude_month(raw_df)
    raw_df = expand_set_items(raw_df, set_config)

    main_df = aggregate_and_sort(raw_df)

    if "_is_set_expanded" in raw_df.columns:
        set_flags = raw_df.groupby(
            raw_df.apply(lambda r: (r["품목명"], r.get("중량",""), r.get("옵션","")), axis=1)
        )["_is_set_expanded"].first()
        def get_flag(r):
            try: return set_flags[(r["품목명"], r["중량"], r["옵션"])]
            except: return False
        main_df["_is_set_expanded"] = main_df.apply(get_flag, axis=1)

    main_df["상품코드"] = main_df.apply(lambda r: match_code(r, code_df), axis=1)
    sheet2_df = build_sheet2(main_df)

    wb  = Workbook()
    ws1 = wb.active; ws1.title = "주문취합"
    apply_style(ws1, main_df)
    ws2 = wb.create_sheet("원두 중량 합산")
    write_simple_sheet(ws2, sheet2_df, ["품목명", "중량(kg)"])
    ws3 = wb.create_sheet("바리스타·농부·농장주")
    write_simple_sheet(ws3, sheet3_df, ["품목명", "빈칸", "이름", "수량"])
    insert_sheet3_into_sheet1(wb)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ══════════════════════════════════════════════
# Streamlit UI
# ══════════════════════════════════════════════
st.set_page_config(page_title="세트 상품 관리", page_icon="📦", layout="wide")

st.markdown("""
<style>
[data-testid="stSidebar"] { background: #FAF3F0; }
h1, h2, h3 { color: #8B3A2A !important; }
.stButton > button {
    background: #8B3A2A; color: white; border: none;
    border-radius: 6px; font-weight: 600;
}
.stButton > button:hover { background: #C4644A; color: white; }
.stDownloadButton > button {
    background: #8B3A2A; color: white; border: none;
    border-radius: 6px; font-weight: 600; width: 100%;
}
</style>
""", unsafe_allow_html=True)

# 상태 초기화
if "set_config" not in st.session_state:
    st.session_state.set_config = load_set_config()
if "editing_set" not in st.session_state:
    st.session_state.editing_set = None

st.title("📦 세트 상품 관리")
st.caption("세트 상품을 구성 품목별로 분리하여 주문취합 엑셀에 반영합니다.")
st.divider()

# ── 좌우 2단 레이아웃 ──
left, right = st.columns([1, 1], gap="large")

# ════════════════════
# 왼쪽: 세트 등록/목록
# ════════════════════
with left:
    st.subheader("세트 상품명")

    # 새 세트 추가
    col_input, col_btn = st.columns([4, 1])
    with col_input:
        new_name = st.text_input("세트 상품명", placeholder="세트 상품명",
                                  label_visibility="collapsed", key="new_set_input")
    with col_btn:
        if st.button("추가", use_container_width=True, key="btn_add"):
            name = new_name.strip()
            if name and name not in st.session_state.set_config:
                st.session_state.set_config[name] = []
                save_set_config(st.session_state.set_config)
                st.session_state.editing_set = name
                st.rerun()
            elif name in st.session_state.set_config:
                st.warning("이미 등록된 세트입니다.")
            else:
                st.warning("세트 상품명을 입력하세요.")

    # 목록 헤더
    h1, h2, h3 = st.columns([5, 3, 2])
    with h1: st.caption("상품명")
    with h2: st.caption("등록일")
    with h3: st.caption("저장")
    st.divider()

    if not st.session_state.set_config:
        st.info("등록된 세트 상품이 없습니다.")
    else:
        for set_name in list(st.session_state.set_config.keys()):
            is_editing = st.session_state.editing_set == set_name
            c1, c2, c3, c4, c5 = st.columns([4, 3, 1, 1, 1])
            with c1:
                st.markdown(f"**{set_name}**")
            with c2:
                st.caption(datetime.today().strftime("%Y-%m-%d"))
            with c3:
                if st.button("선택" if not is_editing else "닫기",
                              key=f"sel_{set_name}", use_container_width=True):
                    st.session_state.editing_set = None if is_editing else set_name
                    st.rerun()
            with c4:
                if st.button("저장", key=f"save_{set_name}", use_container_width=True):
                    save_set_config(st.session_state.set_config)
                    st.success("저장!")
            with c5:
                if st.button("삭제", key=f"del_{set_name}", use_container_width=True):
                    del st.session_state.set_config[set_name]
                    if st.session_state.editing_set == set_name:
                        st.session_state.editing_set = None
                    save_set_config(st.session_state.set_config)
                    st.rerun()

# ════════════════════
# 오른쪽: 구성 옵션 편집
# ════════════════════
with right:
    sel = st.session_state.editing_set
    if not sel or sel not in st.session_state.set_config:
        st.subheader("구성 옵션")
        st.info("← 왼쪽에서 세트 상품을 선택하세요.")
    else:
        st.subheader(f"구성 옵션 — {sel}")
        comps = st.session_state.set_config[sel]

        # 헤더
        hc1, hc2, hc3, hc4, hc5 = st.columns([1, 4, 2, 2, 1])
        with hc1: st.caption("순서")
        with hc2: st.caption("구성 옵션명")
        with hc3: st.caption("등록일")
        with hc4: st.caption("저장")
        st.divider()

        # 기존 구성 목록
        for i, comp in enumerate(comps):
            rc1, rc2, rc3, rc4, rc5 = st.columns([1, 4, 2, 2, 1])
            with rc1:
                new_qty = st.number_input("수량", min_value=1, value=comp.get("qty", 1),
                                           key=f"qty_{sel}_{i}", label_visibility="collapsed")
            with rc2:
                new_nm = st.text_input("품목명", value=comp.get("name", ""),
                                        key=f"nm_{sel}_{i}", label_visibility="collapsed")
            with rc3:
                st.caption(datetime.today().strftime("%Y-%m-%d"))
            with rc4:
                if st.button("저장", key=f"csave_{sel}_{i}", use_container_width=True):
                    comps[i] = {"name": new_nm, "qty": int(new_qty),
                                 "weight": comp.get("weight",""), "option": comp.get("option","")}
                    save_set_config(st.session_state.set_config)
                    st.success("저장!")
            with rc5:
                if st.button("삭제", key=f"cdel_{sel}_{i}", use_container_width=True):
                    comps.pop(i)
                    save_set_config(st.session_state.set_config)
                    st.rerun()

        # 새 구성 품목 추가
        st.divider()
        na1, na2, na3 = st.columns([1, 5, 1])
        with na1:
            add_qty = st.number_input("수량", min_value=1, value=1,
                                       key=f"addqty_{sel}", label_visibility="collapsed")
        with na2:
            add_nm = st.text_input("품목명", placeholder="구성 옵션명",
                                    key=f"addnm_{sel}", label_visibility="collapsed")
        with na3:
            if st.button("추가", key=f"addcomp_{sel}", use_container_width=True):
                if add_nm.strip():
                    comps.append({"name": add_nm.strip(), "qty": int(add_qty),
                                   "weight": "", "option": ""})
                    save_set_config(st.session_state.set_config)
                    st.rerun()

st.divider()

# ════════════════════
# 주문 처리 섹션
# ════════════════════
st.subheader("주문취합 처리")
col1, col2 = st.columns(2)
with col1:
    order_file = st.file_uploader("📄 주문취합 Excel", type=["xlsx"])
with col2:
    code_file  = st.file_uploader("📋 자사몰 상품코드 Excel", type=["xlsx"])

if order_file and code_file:
    if st.button("🚀 처리 시작", use_container_width=True):
        with st.spinner("처리 중..."):
            try:
                buf = process(order_file, code_file, st.session_state.set_config)
                today = datetime.today().strftime("%Y%m%d")
                st.success("✅ 처리 완료!")
                if st.session_state.set_config:
                    st.info(f"📦 세트 분리 적용: {len(st.session_state.set_config)}개 → 하늘색 행으로 표시됨")
                st.download_button(
                    label="⬇️ 결과 Excel 다운로드",
                    data=buf,
                    file_name=f"자사몰주문취합_{today}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"❌ 오류: {e}")
                st.exception(e)
else:
    st.info("파일 두 개를 모두 업로드하면 처리 버튼이 활성화됩니다.")
